#!/usr/bin/env python3
"""
Script to bulk import products from Excel (.xlsx) with automatic base64 image conversion
Excel Format: name, name_en, category_id, description, base_price, image_path, is_hot, is_cold, is_available...

Usage: python scripts/bulk_import_products_excel.py products.xlsx
"""
import sys
import os
import base64
from pathlib import Path

try:
    from openpyxl import load_workbook, Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Add parent directory to path to import models
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from utils.database import db


def image_to_base64(image_path: str) -> str:
    """Convert image file to base64 string"""
    if not image_path or not os.path.exists(image_path):
        return None

    try:
        with open(image_path, 'rb') as image_file:
            image_data = image_file.read()
            base64_string = base64.b64encode(image_data).decode('utf-8')

            # Get file extension to add proper data URI prefix
            ext = Path(image_path).suffix.lower()
            mime_types = {
                '.jpg': 'image/jpeg',
                '.jpeg': 'image/jpeg',
                '.png': 'image/png',
                '.gif': 'image/gif',
                '.webp': 'image/webp',
                '.svg': 'image/svg+xml'
            }

            mime_type = mime_types.get(ext, 'image/jpeg')
            return f"data:{mime_type};base64,{base64_string}"
    except Exception as e:
        print(f"  ⚠️  Lỗi đọc ảnh '{image_path}': {e}")
        return None


def str_to_bool(value):
    """Convert string to boolean"""
    if value is None or value == '':
        return False
    if isinstance(value, bool):
        return value
    return str(value).lower() in ['true', '1', 'yes', 'có', 'y', 'x']


def get_cell_value(row, headers, column_name, default=''):
    """Get cell value from row by column name"""
    try:
        if column_name in headers:
            idx = headers.index(column_name)
            value = row[idx]
            return value if value is not None else default
        return default
    except:
        return default


def import_products_from_excel(excel_path: str):
    """Import products from Excel file"""
    if not OPENPYXL_AVAILABLE:
        print("❌ Lỗi: Cần cài đặt thư viện openpyxl")
        print("   Chạy: pip install openpyxl")
        return False

    if not os.path.exists(excel_path):
        print(f"❌ File không tồn tại: {excel_path}")
        return False

    print(f"\n📂 Đang đọc file: {excel_path}")
    print("="*60)

    success_count = 0
    error_count = 0
    total_count = 0

    try:
        # Load workbook
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        
        # Try to find 'Products' sheet, otherwise use active sheet
        if 'Products' in wb.sheetnames:
            ws = wb['Products']
            print(f"📄 Đọc sheet: Products")
        else:
            ws = wb.active
            print(f"📄 Đọc sheet: {ws.title}")

        # Get headers from first row
        headers = []
        for cell in ws[1]:
            # Bỏ qua cell rỗng và chuyển sang string
            if cell.value is not None:
                headers.append(str(cell.value))
            else:
                headers.append('')

        # Lọc bỏ headers rỗng
        headers = [h for h in headers if h]
        
        print(f"📋 Tìm thấy {len(headers)} cột: {', '.join(headers)}")

        # Validate required headers
        required_headers = ['name', 'category_id', 'base_price']
        if not all(h in headers for h in required_headers):
            print(f"❌ Excel thiếu các cột bắt buộc: {', '.join(required_headers)}")
            print(f"   Các cột hiện có: {', '.join(headers)}")
            return False

        print("\n🔄 Bắt đầu import...\n")

        # Process data rows (skip header)
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Skip empty rows
            if not any(row):
                continue

            total_count += 1
            name = get_cell_value(row, headers, 'name', '').strip()

            if not name:
                print(f"[{total_count}] ⚠️  Bỏ qua dòng {row_num}: Không có tên sản phẩm")
                error_count += 1
                continue

            print(f"[{total_count}] {name}...", end=" ")

            try:
                # Extract data from Excel
                category_id = int(get_cell_value(row, headers, 'category_id', 0))
                if category_id == 0:
                    print(f"❌ Thiếu category_id")
                    error_count += 1
                    continue

                name_en = get_cell_value(row, headers, 'name_en', name)
                name_en = str(name_en).strip() if name_en else name
                
                description = get_cell_value(row, headers, 'description', '')
                description = str(description).strip() if description else ''
                
                ingredients = get_cell_value(row, headers, 'ingredients', '')
                ingredients = str(ingredients).strip() if ingredients else ''
                
                base_price = float(get_cell_value(row, headers, 'base_price', 0))

                if base_price == 0:
                    print(f"❌ Thiếu base_price")
                    error_count += 1
                    continue

                # Image conversion
                image_path = get_cell_value(row, headers, 'image_path', '')
                image_path = str(image_path).strip() if image_path else ''
                image_base64 = None
                if image_path:
                    # If relative path, make it relative to Excel file location
                    if not os.path.isabs(image_path):
                        excel_dir = os.path.dirname(os.path.abspath(excel_path))
                        image_path = os.path.join(excel_dir, image_path)

                    image_base64 = image_to_base64(image_path)

                # Calories
                calories_small = int(get_cell_value(row, headers, 'calories_small', 0) or 0)
                calories_medium = int(get_cell_value(row, headers, 'calories_medium', 0) or 0)
                calories_large = int(get_cell_value(row, headers, 'calories_large', 0) or 0)

                # Boolean attributes
                is_hot = str_to_bool(get_cell_value(row, headers, 'is_hot', True))
                is_cold = str_to_bool(get_cell_value(row, headers, 'is_cold', True))
                is_caffeine_free = str_to_bool(get_cell_value(row, headers, 'is_caffeine_free', False))
                is_available = str_to_bool(get_cell_value(row, headers, 'is_available', True))
                is_featured = str_to_bool(get_cell_value(row, headers, 'is_featured', False))
                is_new = str_to_bool(get_cell_value(row, headers, 'is_new', False))
                is_bestseller = str_to_bool(get_cell_value(row, headers, 'is_bestseller', False))
                is_seasonal = str_to_bool(get_cell_value(row, headers, 'is_seasonal', False))

                # Insert to database
                query = """
                    INSERT INTO products
                    (category_id, name, name_en, description, ingredients, image,
                     base_price, calories_small, calories_medium, calories_large,
                     is_hot, is_cold, is_caffeine_free, is_available, is_featured,
                     is_new, is_bestseller, is_seasonal, created_at, updated_at)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
                """

                product_id = db.insert(query, (
                    category_id, name, name_en, description, ingredients, image_base64,
                    base_price, calories_small, calories_medium, calories_large,
                    is_hot, is_cold, is_caffeine_free, is_available, is_featured,
                    is_new, is_bestseller, is_seasonal
                ))

                if product_id:
                    print(f"✅ (ID: {product_id})")
                    success_count += 1
                else:
                    print(f"❌ Lỗi insert")
                    error_count += 1

            except ValueError as e:
                print(f"❌ Giá trị không hợp lệ: {e}")
                error_count += 1
            except Exception as e:
                print(f"❌ Lỗi: {e}")
                error_count += 1

        wb.close()

    except Exception as e:
        print(f"\n❌ Lỗi đọc file Excel: {e}")
        return False

    # Summary
    print("\n" + "="*60)
    print("📊 KẾT QUẢ:")
    print(f"   Tổng số: {total_count}")
    print(f"   ✅ Thành công: {success_count}")
    print(f"   ❌ Lỗi: {error_count}")
    print("="*60)

    return error_count == 0


def create_sample_excel():
    """Create a sample Excel file"""
    if not OPENPYXL_AVAILABLE:
        print("❌ Lỗi: Cần cài đặt thư viện openpyxl")
        print("   Chạy: pip install openpyxl")
        return None

    sample_path = "products_sample.xlsx"

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    # Headers
    headers = [
        'name', 'name_en', 'category_id', 'description', 'ingredients',
        'base_price', 'image_path', 'calories_small', 'calories_medium', 'calories_large',
        'is_hot', 'is_cold', 'is_caffeine_free', 'is_available',
        'is_featured', 'is_new', 'is_bestseller', 'is_seasonal'
    ]
    ws.append(headers)

    # Sample data
    sample_products = [
        ['Cà Phê Đen Đá', 'Black Coffee', 1, 'Cà phê đen truyền thống', 'Cà phê Robusta',
         35000, 'images/black_coffee.jpg', 5, 10, 15, True, True, False, True, False, False, False, False],
        ['Cà Phê Sữa', 'Milk Coffee', 1, 'Cà phê sữa đặc ngọt dịu', 'Cà phê - Sữa đặc',
         40000, 'images/milk_coffee.jpg', 150, 200, 280, True, True, False, True, True, False, True, False],
        ['Trà Sữa Trân Châu', 'Bubble Tea', 2, 'Trà sữa với trân châu đen', 'Trà đen - Sữa - Trân châu',
         55000, 'images/bubble_tea.jpg', 300, 380, 450, False, True, True, True, False, True, True, False],
        ['Bánh Croissant', 'Croissant', 3, 'Bánh sừng bò giòn tan', 'Bột mì - Bơ',
         35000, 'images/croissant.jpg', 280, 280, 280, False, False, False, True, False, False, False, False],
    ]

    for product in sample_products:
        ws.append(product)

    # Auto-adjust column width
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    # Save
    wb.save(sample_path)
    print(f"✅ Đã tạo file Excel mẫu: {sample_path}")
    return sample_path


def main():
    """Main function"""
    if not OPENPYXL_AVAILABLE:
        print("="*60)
        print("❌ THIẾU THƯ VIỆN: openpyxl")
        print("="*60)
        print("\n📦 Cài đặt:")
        print("   pip install openpyxl")
        print("\n   Hoặc:")
        print("   pip3 install openpyxl")
        return

    if len(sys.argv) < 2:
        print("="*60)
        print("📦 BULK IMPORT SẢN PHẨM TỪ EXCEL")
        print("="*60)
        print("\n📖 Cách dùng:")
        print(f"   python {sys.argv[0]} <file.xlsx>")
        print("\n📋 Format Excel (các cột bắt buộc):")
        print("   - name: Tên sản phẩm (Tiếng Việt)")
        print("   - category_id: ID danh mục")
        print("   - base_price: Giá cơ bản (VND)")
        print("\n📋 Format Excel (các cột tùy chọn):")
        print("   - name_en: Tên tiếng Anh")
        print("   - description: Mô tả")
        print("   - ingredients: Thành phần")
        print("   - image_path: Đường dẫn file ảnh")
        print("   - calories_small/medium/large: Calories")
        print("   - is_hot, is_cold, is_available, etc.: TRUE/FALSE")

        print("\n" + "-"*60)
        choice = input("\n💡 Tạo file Excel mẫu? [Y/n]: ").strip().lower()
        if choice in ['', 'y', 'yes', 'có']:
            sample_file = create_sample_excel()
            if sample_file:
                print(f"\n📝 Bạn có thể mở file '{sample_file}' bằng Excel/LibreOffice và chỉnh sửa")
                print(f"   Sau đó chạy lại:")
                print(f"   python {sys.argv[0]} {sample_file}")
        return

    excel_path = sys.argv[1]
    import_products_from_excel(excel_path)


if __name__ == "__main__":
    main()

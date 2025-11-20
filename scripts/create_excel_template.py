#!/usr/bin/env python3
"""
Create Excel template for product import
"""
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("❌ Cần cài đặt: pip install openpyxl")
    exit(1)

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Products"

# Headers with formatting
headers = [
    'name', 'name_en', 'category_id', 'description', 'ingredients',
    'base_price', 'image_path', 'calories_small', 'calories_medium', 'calories_large',
    'is_hot', 'is_cold', 'is_caffeine_free', 'is_available',
    'is_featured', 'is_new', 'is_bestseller', 'is_seasonal'
]

# Add headers with styling
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center")

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment

# Sample data (Vietnamese products)
sample_products = [
    # Cà phê (category_id = 1)
    ['Phin Cà Phê Sữa Đá', 'Iced Milk Coffee', 1, 'Cà phê phin truyền thống kết hợp sữa đặc', 'Cà phê Robusta + Arabica - Sữa đặc',
     45000, '', 150, 200, 280, True, True, False, True, True, True, True, False],
    ['Bạc Xỉu', 'Bac Xiu', 1, 'Cà phê sữa nhẹ nhàng ngọt dịu', 'Cà phê - Sữa tươi',
     45000, '', 180, 230, 300, True, True, False, True, False, False, False, False],
    ['Americano', 'Americano', 1, 'Cà phê đen pha espresso', 'Espresso',
     40000, '', 10, 15, 20, True, True, False, True, False, False, False, False],
    ['Cappuccino', 'Cappuccino', 1, 'Espresso với sữa tươi và foam mịn', 'Espresso - Sữa tươi',
     55000, '', 120, 160, 220, True, True, False, True, False, False, True, False],
    ['Latte', 'Latte', 1, 'Espresso với nhiều sữa tươi', 'Espresso - Sữa tươi',
     55000, '', 150, 200, 250, True, True, False, True, False, False, False, False],

    # Trà (category_id = 2)
    ['Trà Sữa Trân Châu Đường Đen', 'Brown Sugar Milk Tea', 2, 'Trà sữa kết hợp trân châu và đường đen', 'Trà đen - Sữa tươi - Trân châu - Đường đen',
     55000, '', 300, 380, 450, False, True, True, True, False, True, True, False],
    ['Trà Đào Cam Sả', 'Peach Passion Fruit Tea', 2, 'Trà trái cây tươi mát', 'Trà xanh - Đào - Cam - Sả',
     50000, '', 120, 180, 240, False, True, True, True, False, False, False, True],
    ['Trà Sữa Matcha', 'Matcha Milk Tea', 2, 'Trà sữa matcha Nhật Bản', 'Matcha - Sữa tươi',
     60000, '', 200, 280, 350, False, True, True, True, True, False, False, False],

    # Bánh ngọt (category_id = 3)
    ['Bánh Croissant Bơ', 'Butter Croissant', 3, 'Bánh sừng bò giòn tan thơm bơ', 'Bột mì - Bơ',
     35000, '', 280, 280, 280, False, False, False, True, False, False, False, False],
    ['Tiramisu', 'Tiramisu', 3, 'Bánh Tiramisu truyền thống Ý', 'Mascarpone - Cà phê - Bánh ladyfinger',
     50000, '', 350, 350, 350, False, False, False, True, False, False, False, False],
    ['Bánh Mì Que Phô Mai', 'Cheese Breadstick', 3, 'Bánh mì que giòn với phô mai', 'Bột mì - Phô mai',
     25000, '', 200, 200, 200, False, False, False, True, False, False, False, False],

    # Sinh tố (category_id = 4)
    ['Sinh Tố Bơ', 'Avocado Smoothie', 4, 'Sinh tố bơ béo ngậy', 'Bơ - Sữa đặc - Đá',
     45000, '', 250, 320, 400, False, True, False, True, False, False, False, False],
    ['Sinh Tố Dâu', 'Strawberry Smoothie', 4, 'Sinh tố dâu tây tươi', 'Dâu tây - Sữa chua - Đường',
     45000, '', 180, 250, 320, False, True, False, True, False, False, False, True],
    ['Sinh Tố Xoài', 'Mango Smoothie', 4, 'Sinh tố xoài nhiệt đới', 'Xoài - Sữa tươi - Đá',
     45000, '', 200, 280, 350, False, True, False, True, False, False, False, True],

    # Đá xay (category_id = 5)
    ['Đá Xay Socola', 'Chocolate Frappe', 5, 'Đồ uống đá xay socola đậm đà', 'Socola - Sữa tươi - Kem - Đá xay',
     60000, '', 350, 450, 550, False, True, False, True, False, False, False, False],
    ['Đá Xay Caramel', 'Caramel Frappe', 5, 'Đá xay caramel ngọt ngào', 'Caramel - Sữa tươi - Kem - Đá xay',
     60000, '', 350, 450, 550, False, True, False, True, False, False, False, False],
]

# Add sample data
for product in sample_products:
    ws.append(product)

# Auto-adjust column width
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if cell.value and len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 50)  # Max 50 characters
    ws.column_dimensions[column_letter].width = adjusted_width

# Freeze header row
ws.freeze_panes = "A2"

# Add instructions sheet
ws_instructions = wb.create_sheet("Hướng Dẫn")
instructions = [
    ["🍵 HƯỚNG DẪN IMPORT SẢN PHẨM TỪ EXCEL", ""],
    ["", ""],
    ["📋 CÁC CỘT BẮT BUỘC:", ""],
    ["name", "Tên sản phẩm (Tiếng Việt)"],
    ["category_id", "ID danh mục (1=Cà phê, 2=Trà, 3=Bánh, 4=Sinh tố, 5=Đá xay)"],
    ["base_price", "Giá cơ bản (VND)"],
    ["", ""],
    ["📋 CÁC CỘT TÙY CHỌN:", ""],
    ["name_en", "Tên tiếng Anh"],
    ["description", "Mô tả sản phẩm"],
    ["ingredients", "Thành phần"],
    ["image_path", "Đường dẫn file ảnh (tương đối hoặc tuyệt đối)"],
    ["calories_small", "Calories size S"],
    ["calories_medium", "Calories size M"],
    ["calories_large", "Calories size L"],
    ["is_hot", "Có phục vụ nóng? (TRUE/FALSE)"],
    ["is_cold", "Có phục vụ lạnh? (TRUE/FALSE)"],
    ["is_caffeine_free", "Không caffeine? (TRUE/FALSE)"],
    ["is_available", "Đang bán? (TRUE/FALSE)"],
    ["is_featured", "Sản phẩm nổi bật? (TRUE/FALSE)"],
    ["is_new", "Sản phẩm mới? (TRUE/FALSE)"],
    ["is_bestseller", "Bán chạy? (TRUE/FALSE)"],
    ["is_seasonal", "Theo mùa? (TRUE/FALSE)"],
    ["", ""],
    ["🚀 CÁCH SỬ DỤNG:", ""],
    ["1.", "Chỉnh sửa sheet 'Products'"],
    ["2.", "Thêm/sửa/xóa sản phẩm theo ý muốn"],
    ["3.", "Lưu file"],
    ["4.", "Chạy: python scripts/bulk_import_products_excel.py <file.xlsx>"],
    ["", ""],
    ["💡 LƯU Ý:", ""],
    ["", "- Không xóa dòng header (dòng đầu tiên)"],
    ["", "- Category ID phải tồn tại trong database"],
    ["", "- Đường dẫn ảnh có thể để trống nếu không có ảnh"],
    ["", "- TRUE/FALSE không phân biệt hoa thường"],
]

for row in instructions:
    ws_instructions.append(row)

# Format instructions
ws_instructions.column_dimensions['A'].width = 20
ws_instructions.column_dimensions['B'].width = 60

# Save
output_file = "products_template.xlsx"
wb.save(output_file)
print(f"✅ Đã tạo file Excel template: {output_file}")
print(f"   - Sheet 'Products': 16 sản phẩm mẫu")
print(f"   - Sheet 'Hướng Dẫn': Hướng dẫn sử dụng")
print(f"\n📝 Mở file bằng Excel/LibreOffice để chỉnh sửa")
print(f"   Sau đó chạy: python scripts/bulk_import_products_excel.py {output_file}")
